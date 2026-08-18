"""
Regenerate /data/harga-produk.json -- daftar produk untuk fitur Survei Harga.

Gabungan dari:
  - data/sku-lmt-spm.json (SKU wajib LMT SPM -- untuk 8 toko EC BIG, flag dipertahankan)
  - data/sku-haba-dt.json (SKU wajib HABA DT -- untuk toko Beauty/Cosmetic Expert
    Traditional: SAGA BEAUTY, DEDE MAMA, flag dipertahankan)
  - Sheet "Detail Mekanisme" di file Consumer Promo (kolom SKU CODE MAPPING,
    SKU DESCRIPTION CURRENT, RSP) -- SKU yang cuma ada di sini diberi flag None.

Kalau ada file Consumer Promo baru (misal untuk bulan berikutnya), jalankan:
  python3 scripts/build_harga_produk.py "Consumer_Promo_xxx.xlsx"

Membutuhkan: pip install openpyxl --break-system-packages
"""
import sys
import json
from pathlib import Path
import openpyxl

DATA_DIR = Path("data")
WAJIB_FILES = ['sku-lmt-spm.json', 'sku-haba-dt.json']


def main():
    if len(sys.argv) < 2:
        print("Pakai: python3 scripts/build_harga_produk.py <file_consumer_promo.xlsx>")
        sys.exit(1)
    promo_path = sys.argv[1]

    wajib_by_pcode = {}
    for fname in WAJIB_FILES:
        items = json.loads((DATA_DIR / fname).read_text())
        for it in items:
            if it['pcode']:
                wajib_by_pcode[it['pcode']] = it

    wb = openpyxl.load_workbook(promo_path, data_only=True)
    ws = wb['Detail Mekanisme']
    # Baris 1-2 header, data mulai baris 3. Kolom: B=BARCODE, J=SKU CODE MAPPING,
    # K=SKU DESCRIPTION CURRENT, N=RSP (sesuai struktur file Agustus 2026 --
    # cek ulang urutan kolom kalau template berubah di file berikutnya).
    promo = {}
    for r in range(3, ws.max_row + 1):
        barcode = ws.cell(row=r, column=2).value
        pcode = ws.cell(row=r, column=10).value
        name = ws.cell(row=r, column=11).value
        rsp = ws.cell(row=r, column=14).value
        if not pcode:
            continue
        pcode = str(pcode).strip()
        promo[pcode] = {
            'barcode': str(barcode).strip() if barcode else None,
            'name': str(name).strip() if name else '',
            'rsp': rsp,
        }

    merged = {}
    for pcode, it in wajib_by_pcode.items():
        merged[pcode] = {
            'pcode': pcode,
            'barcode': it['barcode'],
            'name': it['name'],
            'flag': it['flag'],
            'isi': it['isi'],
            'rsp': promo.get(pcode, {}).get('rsp'),
            'source': 'wajib',
        }
    for pcode, it in promo.items():
        if pcode in merged:
            merged[pcode]['rsp'] = it['rsp']
            continue
        merged[pcode] = {
            'pcode': pcode,
            'barcode': it['barcode'],
            'name': it['name'],
            'flag': None,
            'isi': None,
            'rsp': it['rsp'],
            'source': 'promo',
        }

    result = sorted(merged.values(), key=lambda x: x['name'])
    (DATA_DIR / 'harga-produk.json').write_text(json.dumps(result, ensure_ascii=False, indent=2))
    print(f"harga-produk.json: {len(result)} produk "
          f"({sum(1 for x in result if x['source']=='wajib')} dari SKU wajib, "
          f"{sum(1 for x in result if x['source']=='promo')} cuma dari file promo)")


if __name__ == '__main__':
    main()
