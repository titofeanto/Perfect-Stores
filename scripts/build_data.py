"""
Regenerate /data/*.json from the source Excel files.

Sumber data (taruh di folder yang sama saat menjalankan script ini, atau
sesuaikan path di bawah):
  - COTC_BY_GROUP.xlsx   -> sheet "LIST CUSTOMER" (daftar toko + area + scope channel)
                             + sheet "LMT SPM" / "LOCAL MINIS" / "HABA DT" (SKU wajib per grup)
  - PH_SKU_AUDIT.xlsx    -> master data produk nasional, dipakai untuk mencari
                             "PC Code" / Parent SKU dari tiap barcode (dipakai untuk
                             mencocokkan dengan PCODE di laporan stock distributor).

Jalankan: python3 scripts/build_data.py
Membutuhkan: pip install openpyxl --break-system-packages
"""
import json
import re
from pathlib import Path
import openpyxl

SOURCE_DIR = Path(".")          # ganti kalau file excel-nya ada di folder lain
OUTPUT_DIR = Path("data")

SCOPE_SLUG_MAP = {
    'LMT SPM': 'lmt-spm',
    'LOCAL MINIS': 'local-minis',
    'HABA DT': 'haba-dt',
}
SHEET_TO_OUTPUT = {
    'LMT SPM': 'sku-lmt-spm',
    'LOCAL MINIS': 'sku-local-minis',
    'HABA DT': 'sku-haba-dt',
}


def build_barcode_to_pcode_lookup(master_path):
    """Returns {barcode: {'pcode': str, 'isi': int|None}}.
    'isi' = jumlah pcs per karton, dari kolom 'PC/CS' di PH_SKU_AUDIT.xlsx
    (kolom ini ditambahkan belakangan oleh tim -- kalau file master Anda versi
    lama belum punya kolom ini, 'isi' akan selalu None dan Karton tidak bisa
    dikonversi otomatis untuk SKU itu)."""
    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb['Sheet1']
    lookup = {}
    for r in range(2, ws.max_row + 1):
        pc_cs = ws.cell(row=r, column=3).value          # kolom C: PC/CS (isi per karton)
        parent_sku = ws.cell(row=r, column=9).value      # kolom I: Parent SKU
        parent_barcode = ws.cell(row=r, column=11).value  # kolom K: Parent Product Barcode
        if parent_barcode is not None and parent_sku is not None:
            try:
                isi = int(pc_cs) if pc_cs is not None else None
            except (ValueError, TypeError):
                isi = None
            lookup[str(parent_barcode).strip()] = {'pcode': str(parent_sku).strip(), 'isi': isi}
    return lookup


def build_stores(group_path):
    wb = openpyxl.load_workbook(group_path, data_only=True)
    ws = wb['LIST CUSTOMER']
    stores = []
    # Kode pelanggan bisa berupa satu blok angka ("941203...689") ATAU beberapa
    # blok angka dipisah tanda hubung ("941203...865740-2" untuk cabang/sub-lokasi).
    # Regex ini menelan SEMUA blok angka berturutan (dipisah "-") sebagai id,
    # baru setelah itu nama toko dimulai -- supaya "kode-2", "kode-3" dst tidak
    # ketiban dianggap toko yang sama dengan "kode" polos.
    id_pattern = re.compile(r'^((?:\d+-)*\d+)-(.+)$')
    for r in range(2, ws.max_row + 1):
        area = ws.cell(row=r, column=1).value
        scope = ws.cell(row=r, column=2).value
        sub = ws.cell(row=r, column=3).value
        name = ws.cell(row=r, column=4).value
        if not area:
            continue
        m = id_pattern.match(str(name).strip())
        store_id, store_name = (m.group(1), m.group(2).strip()) if m else (str(name), str(name))
        stores.append({
            'id': store_id,
            'name': store_name,
            'area': area,
            'scopeChannel': scope,
            'scopeSlug': SCOPE_SLUG_MAP.get(scope, str(scope).lower().replace(' ', '-')),
            'subChannel': sub,
        })
    return stores


def build_sku_lists(group_path, barcode_to_info):
    wb = openpyxl.load_workbook(group_path, data_only=True)
    results = {}
    for sheet_name, out_name in SHEET_TO_OUTPUT.items():
        ws = wb[sheet_name]
        items = []
        for r in range(2, ws.max_row + 1):
            barcode = ws.cell(row=r, column=1).value
            prod = ws.cell(row=r, column=2).value
            ket = ws.cell(row=r, column=3).value
            if barcode is None:
                continue
            bstr = str(barcode).strip()
            info = barcode_to_info.get(bstr, {})
            items.append({
                'barcode': bstr,
                'pcode': info.get('pcode'),  # None kalau tidak ketemu di master data
                'isi': info.get('isi'),      # jumlah pcs per karton, None kalau tidak tersedia
                'name': str(prod).strip() if prod else '',
                'flag': str(ket).strip() if ket else 'COTC',
            })
        results[out_name] = items
    return results


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)
    barcode_to_info = build_barcode_to_pcode_lookup(SOURCE_DIR / 'PH_SKU_AUDIT.xlsx')

    stores = build_stores(SOURCE_DIR / 'COTC_BY_GROUP.xlsx')
    (OUTPUT_DIR / 'stores.json').write_text(json.dumps(stores, ensure_ascii=False, indent=2))
    print(f"stores.json: {len(stores)} toko")

    sku_lists = build_sku_lists(SOURCE_DIR / 'COTC_BY_GROUP.xlsx', barcode_to_info)
    for out_name, items in sku_lists.items():
        path = OUTPUT_DIR / f'{out_name}.json'
        path.write_text(json.dumps(items, ensure_ascii=False, indent=2))
        no_pcode = sum(1 for i in items if not i['pcode'])
        no_isi = sum(1 for i in items if not i['isi'])
        print(f"{out_name}.json: {len(items)} SKU ({no_pcode} tanpa PC Code, {no_isi} tanpa isi/karton)")


if __name__ == '__main__':
    main()
